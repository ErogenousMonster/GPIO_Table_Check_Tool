# coding=utf-8
# 2019.07.10 从文件中提取所需数据
# python 3.7

# 2019.07.18 增加对同时拥有三个属性值的bom “I,NI,NI”的情况进行判断
# 三种情况分别表示三种不同类型的板子，只是EE为了方便写在一起
# 解决方法直接生成：第一版，输入一个sheet，生成三份sheet；后续版本，输入三个sheet，生成三个sheet
# 三个sheet的命名由EE决定，暂时定为 BOM1,BOM2,BOM3

# 2019.07.18 增加对NI，I的分类判断：如果PU_PD显示有reserved字样，则需要将NI视为通路，并需验证reserved属于NI
# 2019.08.08 后续版本的添加（可以跑出多个版本）
# 2019.08.27 忽略AIO的项（只show出信息不做判断），并更改比较数据
# 2019.09.03 一律判斷BOM為 I 的PU或PD的個數是否超過兩個（兩個以上PU 或者兩個以上PD），不用去管PU/PD項是否包含Reserved字樣。

# User Guide
# 输入要求：
# 1. 初始输入只能有一个excel，内含两个sheet，"Power Rail"（名称确定）和其他，"Power Rail"第一列为"Power net name",
# 第二列为"Common power name"
# 2. 初始输入的 GPIO sheet需要用 Group（不区分大小写）分割开，顺序一定要为pin name, signal name, pu/pd,
# resistance, voltage, note,并且 pin name 为第一列（前面不能包含其他列）
# 3. 如果 pin name有拼写错误，会生成 error excel。
# 4. signal name 会匹配出pin及出pin的第二条线，如果匹配，则填入匹配的那条线，如果不匹配则填入出pin端的线


import os
import re
import xlwings as xw
import copy
from math import ceil
import xlsxwriter
import openpyxl
import xlrd


# 定义错误输出
def create_error_message(excel_path, error_message):
    # 創建excel
    workbook = xlsxwriter.Workbook(excel_path)
    worksheet = workbook.add_worksheet('error_message')

    title_format = workbook.add_format({'font_size': 22})
    error_format = workbook.add_format({'font_size': 18})

    if isinstance(error_message, str):
        worksheet.write('A1', 'Program running error:', title_format)
        worksheet.write('B2', error_message + ', please check and try again!', error_format)
    if isinstance(error_message, list):
        worksheet.write('A1', 'Error pins:', title_format)
        worksheet.write_column('B2', error_message, error_format)

    workbook.close()
    raise FileNotFoundError


# 将多维list展开成一维
def flatten(a):
    if not isinstance(a, (list,)) and not isinstance(a, (tuple,)):
        return [a]
    else:
        b = []
        for item in a:
            b += flatten(item)
    return b


# 自适应表格
def sheets_autofit(excel_path):
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(excel_path)
    sheets = wb.sheets
    for active_sheet in sheets:
        active_sheet.autofit('c')

    wb.save()
    app.quit()


def compare_with_common_design(result_flag, real_data_list, common_design_list,
                               power_common_power_dict, all_list, gpp_idx):
    """用于比较当前版本与common design的差异并得出正误"""
    half_result_list = []
    check_net, check_pu_pd, check_resistance, check_voltage = common_design_list
    check_pu_pd = check_pu_pd.replace(' ', '') if check_pu_pd else None
    check_resistance = check_resistance.replace(' ', '') if check_resistance else None
    check_voltage = check_voltage.replace(' ', '') if check_voltage else None
    real_net, real_pu_pd, real_resistance, real_voltage = real_data_list
    real_signal_name_list, real_pu_pd_list, real_resistance_list, real_power_list = all_list
    real_pu_pd = real_pu_pd.replace(' ', '') if real_pu_pd else None
    real_resistance = real_resistance.replace(' ', '') if real_resistance else None
    real_voltage = real_voltage.replace(' ', '') if real_voltage else None
    # 对net进行判断
    if check_net:
        # 对线进行判断
        # 如果net是reserved则判断实际net是否存在
        if check_net == 'RESERVED':
            if real_net == ['NC']:
                half_result_list.append(0)
            else:
                # 选择出PIN端的线
                if real_net and len(real_net) > 1:
                    real_signal_name_list[gpp_idx] = [real_net[0]]
                half_result_list.append(1)
        # 如果存在明确的线，就比较明确的线
        else:
            # 如果线匹配，则存匹配的线
            if check_net.strip() in real_net:
                real_signal_name_list[gpp_idx] = [check_net.strip()]
                half_result_list.append(1)
            else:
                # 如果线不匹配，则选择出PIN端的线
                if real_net and len(real_net) > 1:
                    real_signal_name_list[gpp_idx] = [real_net[0]]
                half_result_list.append(0)
    # 如果没有check net
    else:
        if real_net == ['NC']:
            half_result_list.append(1)
        # 不存在show false
        else:
            if real_net and len(real_net) > 1:
                real_signal_name_list[gpp_idx] = [real_net[0]]
            half_result_list.append(0)
    # 如果result_flag 为True说明不存在有BOM为I的两个PU或PD的情况

    if result_flag:
        # 对pupd进行判断
        pu_pd_flag = False
        if real_pu_pd:
            real_pupd_list = real_pu_pd.upper().split('/')
        else:
            real_pupd_list = []

        if check_pu_pd:
            check_pu_pd = check_pu_pd.upper()
            check_pupd_list = check_pu_pd.split('/')

            # 如果check_pu_pd中不存在reserved说明不需要考虑NI的情况，则在真实数据删除
            if check_pu_pd.find('RESERVED') == -1:
                reserved_idx_list = [idx for idx in range(len(real_pupd_list))
                                     if real_pupd_list[idx].find('RESERVED') == -1]
                real_pu_pd_list[-1] = '/'.join([real_pupd_list[idx] for idx in reserved_idx_list])
                real_resistance_list[-1] = '/'.join([real_resistance_list[-1].split('/')[idx]
                                                     for idx in reserved_idx_list])
                real_power_list[-1] = '/'.join([real_power_list[-1].split('/')[idx]
                                                for idx in reserved_idx_list])
        else:
            # 如果check_pu_pd为空，也不需要考虑NI的情况，则在真实数据删除
            reserved_idx_list = [idx for idx in range(len(real_pupd_list))
                                 if real_pupd_list[idx].find('RESERVED') == -1]
            real_pu_pd_list[-1] = '/'.join([real_pupd_list[idx] for idx in reserved_idx_list])
            real_resistance_list[-1] = '/'.join([real_resistance_list[-1].split('/')[idx]
                                                 for idx in reserved_idx_list])
            real_power_list[-1] = '/'.join([real_power_list[-1].split('/')[idx]
                                            for idx in reserved_idx_list])
            check_pupd_list = []

        real_pu_pd, real_resistance, real_voltage = real_pu_pd_list[-1], \
                                                    real_resistance_list[-1], real_power_list[-1]
        real_pu_pd = real_pu_pd.replace(' ', '') if real_pu_pd else None
        real_resistance = real_resistance.replace(' ', '') if real_resistance else None
        real_voltage = real_voltage.replace(' ', '') if real_voltage else None
        # 将real_voltage中的GND去掉
        real_voltage = '/'.join([item for item in real_voltage.split('/') if item.find('GND') == -1])\
            if real_voltage else None
        if real_pu_pd:
            real_pupd_list = real_pu_pd.upper().split('/')
        else:
            real_pupd_list = []

        # 如果两个不同则报错
        if sorted(check_pupd_list) != sorted(real_pupd_list):
            half_result_list.append(0)
        else:
            half_result_list.append(1)
            pu_pd_flag = True

        # 对resistance进行判断
        resistance_flag = False
        if check_resistance:
            check_res_list = check_resistance.split('/')
            check_res_list = [x.strip() for x in check_res_list]
        else:
            check_res_list = []
        if real_resistance:
            real_res_list = real_resistance.split('/')
        else:
            real_res_list = []

        # 如果两个不同则报错
        if sorted(check_res_list) != sorted(real_res_list):
            half_result_list.append(0)
        else:
            half_result_list.append(1)
            resistance_flag = True

        # 对voltage进行判断
        voltage_flag = False
        if check_voltage:
            check_voltage_list = check_voltage.split('/')
            check_voltage_list = [x.strip() for x in check_voltage_list]
        else:
            check_voltage_list = []
        if real_voltage:
            real_voltage_list = real_voltage.split('/')
            real_voltage_list = [power_common_power_dict.get(x, x) for x in real_voltage_list]
        else:
            real_voltage_list = []

        # 如果两个不同则报错
        if sorted(check_voltage_list) != sorted(real_voltage_list):
            half_result_list.append(0)
        else:
            half_result_list.append(1)
            voltage_flag = True

        # 如果全部相同并且不为空，则比较顺序是否相同
        if pu_pd_flag and resistance_flag and voltage_flag and check_pupd_list:
            half_result_list = half_result_list[:-3]
            for idx in range(len(check_pupd_list)):
                if check_pupd_list[idx].find('PD') > -1:
                    check_voltage_list.insert(idx, 'GND')

            for idx in range(len(real_pupd_list)):
                if real_pupd_list[idx].find('PD') > -1:
                    real_voltage_list.insert(idx, 'GND')

            check_all_list = [[check_pupd_list[pud_ind], check_res_list[pud_ind],
                               check_voltage_list[pud_ind]]
                              for pud_ind in range(len(check_pupd_list))]

            real_all_list = [[real_pupd_list[pud_ind], real_res_list[pud_ind],
                              real_voltage_list[pud_ind]]
                             for pud_ind in range(len(real_pupd_list))]

            if sorted(check_all_list) == sorted(real_all_list):
                half_result_list += [1, 1, 1]
            else:
                half_result_list += [0, 0, 0]

    else:
        half_result_list += [0, 0, 0]
    half_result_list = [str(x) for x in half_result_list]
    return half_result_list


def compare_with_previous_design(real_data_list, previous_design_list, real_signal_name_list, gpp_idx):
    """用于比较当前版本与previous design的差异并得出正误"""
    half_result_list = []
    previous_net, previous_pu_pd, previous_resistance, previous_voltage = previous_design_list
    previous_pu_pd = previous_pu_pd.replace(' ', '') if previous_pu_pd else None
    previous_resistance = previous_resistance.replace(' ', '') if previous_resistance else None
    previous_voltage = previous_voltage.replace(' ', '') if previous_voltage else None
    real_net, real_pu_pd, real_resistance, real_voltage = real_data_list
    real_pu_pd = real_pu_pd.replace(' ', '') if real_pu_pd else None
    real_resistance = real_resistance.replace(' ', '') if real_resistance else None
    real_voltage = real_voltage.replace(' ', '') if real_voltage else None
    # 将real_voltage中的GND去掉
    if real_voltage:
        real_voltage = '/'.join([item for item in real_voltage.split('/') if item.find('GND') == -1])
    # common design的比较方式与previous的比较方式不同
    # 1.Signal name是直接进行比较
    # 如果有多条线，进行判断
    if len(real_net) > 1:
        # 选择出PIN端的线
        # 如果线匹配，则存匹配的线
        if previous_net in real_net:
            real_signal_name_list[gpp_idx] = [previous_net]
            half_result_list.append(1)
        else:
            # 如果线不匹配，则选择出PIN端的线
            real_signal_name_list[gpp_idx] = [real_net[0]]
            half_result_list.append(0)
    # 如果只有一条线，则直接进行比较
    else:
        if real_net[0] == previous_net:
            half_result_list.append(1)
            real_signal_name_list[gpp_idx] = [previous_net]
        else:
            real_signal_name_list[gpp_idx] = [real_net[0]]
            half_result_list.append(0)

    # 2.pu/pd比较
    # 如果两个都有值，进行比较
    if previous_pu_pd and real_pu_pd:
        previous_pu_pd = previous_pu_pd.upper()
        previous_pu_pd_list = previous_pu_pd.split('/')
        real_pupd_list = real_pu_pd.upper().split('/')
        previous_resistance_list = previous_resistance.split('/')
        real_resistance_list = real_resistance.split('/')
        try:
            previous_voltage_list = previous_voltage.split('/')
        except:
            previous_voltage_list = ['']

        real_voltage_list = real_voltage.split('/') if real_voltage else ['']

        for pud_idx in range(len(previous_pu_pd_list)):
            previous_pu_pd_item = previous_pu_pd_list[pud_idx]
            try:
                real_pu_pd_item = real_pupd_list[pud_idx]
            except:
                real_pu_pd_item = ''

            if previous_pu_pd_item.find('PD') > -1:
                previous_voltage_list.insert(pud_idx, 'GND')
            if real_pu_pd_item.find('PD') > -1:
                real_voltage_list.insert(pud_idx, 'GND')

        # 先判断是否是个数错误，如果不是则去判断是否是顺序错误
        if sorted(previous_pu_pd_list) == sorted(real_pupd_list) \
                and sorted(previous_resistance_list) == sorted(real_resistance_list) \
                and sorted(previous_voltage_list) == sorted(real_voltage_list):

            previous_all_list = [[previous_pu_pd_list[pud_ind], previous_resistance_list[pud_ind],
                                  previous_voltage_list[pud_ind]]
                                 for pud_ind in range(len(previous_pu_pd_list))]

            real_all_list = [[real_pupd_list[pud_ind], real_resistance_list[pud_ind],
                              real_voltage_list[pud_ind]]
                             for pud_ind in range(len(real_pupd_list))]

            if sorted(previous_all_list) == sorted(real_all_list):
                half_result_list += [1, 1, 1]
            else:
                half_result_list += [0, 0, 0]

        # 个数错误去找是哪个个数错误
        else:
            if sorted(previous_pu_pd_list) == sorted(real_pupd_list):
                half_result_list.append(1)
            else:
                half_result_list.append(0)
            if sorted(previous_resistance_list) == sorted(real_resistance_list):
                half_result_list.append(1)
            else:
                half_result_list.append(0)
            if sorted(previous_voltage_list) == sorted(real_voltage_list):
                half_result_list.append(1)
            else:
                half_result_list.append(0)
    else:
        if previous_pu_pd in [None, ''] and real_pu_pd in [None, '']:
            half_result_list.append(1)
        else:
            half_result_list.append(0)

        if previous_resistance in [None, ''] and real_resistance in [None, '']:
            half_result_list.append(1)
        else:
            half_result_list.append(0)

        if previous_voltage in [None, ''] and real_voltage in [None, '']:
            half_result_list.append(1)
        else:
            half_result_list.append(0)

    return half_result_list


# 输入excel处理器
class InputExcelHandler:
    def __init__(self):
        root_path = os.getcwd()
        self._root_path = os.path.join('\\'.join(root_path.split('\\')[:-1]), 'input')
        self._output_path = os.path.join('\\'.join(root_path.split('\\')[:-1]), 'output')
        self.error_excel_path = os.path.join(self._output_path, 'error.xlsx')

        # get_input_excel_path()
        self.gpio_version_input_excel_path = os.path.join(self._root_path, 'Input_GPIO_Table.xlsx')
        self.gpio_version_input_excel_path = self.gpio_version_input_excel_path if os.path.exists(self.gpio_version_input_excel_path) else None
        self.gpio_common_design_excel_path = None

        # get_version()
        self.first_version_flag = True
        self.app = None
        self.wb = None
        self.common_compare_sht = None
        self.common_sht_list = []

        # commoon desgin compare的数据
        self.origin_common_compare_excel_list = []
        self.last_common_compare_list = []
        self.first_two_line_common_compare_list = []

        # get_data_from_common_design_excel
        self.common_design_pin_list = []
        self.common_design_list = []
        self.common_design_pin_net_dict = {}
        self.common_design_pin_pu_pd_dict = {}
        self.common_design_pin_resistance_dict = {}
        self.common_design_pin_voltage_dict = {}

        self.power_rail_list = []
        self.power_rail_content_list = []
        self.power_common_power_dict = {}
        self.aio_ignore_idx_list = []

        # get_data_from_version_excel
        # 前一版本表格数据
        self.previous_design_pin_net_dict = {}
        self.previous_design_pin_pu_pd_dict = {}
        self.previous_design_pin_resistance_dict = {}
        self.previous_design_pin_voltage_dict = {}
        self.first_line_version_list = []

        # common desgin compare的数据
        self.origin_common_compare_excel_list = []
        self.last_common_compare_list = []
        self.first_two_line_common_compare_list = []
        self.origin_excel_list = []

        self.col_len = None

    def get_input_excel_path(self):
        """获取输入excel的路径"""
        for x in os.listdir(self._root_path):
            (shotname, extension) = os.path.splitext(x)
            if extension == '.xlsx' and shotname.find('$') == -1 and shotname != 'Input_GPIO_Table':
                self.gpio_common_design_excel_path = os.path.join(self._root_path, x)

        # 如果输入没有gpio_common_design说明确实，则报错
        if self.gpio_common_design_excel_path is None:
            error_message = 'GPIO common design excel is not exists.'
            create_error_message(self.error_excel_path, error_message)
        # print('self.gpio_common_design_excel_path', self.gpio_common_design_excel_path)

    def get_data_from_common_design_excel(self):
        """从输入excel中获取所需数据"""
        self.common_design_pin_list = []
        self.common_design_list = []
        self.common_design_pin_net_dict = {}
        self.common_design_pin_pu_pd_dict = {}
        self.common_design_pin_resistance_dict = {}
        self.common_design_pin_voltage_dict = {}

        self.power_rail_list = []
        self.power_rail_content_list = []
        self.power_common_power_dict = {}
        self.aio_ignore_idx_list = []

        self.app = xw.App(visible=False, add_book=False)
        self.wb = self.app.books.open(self.gpio_common_design_excel_path)
        shts = self.wb.sheets
        power_sht = None
        common_sht = None

        for sht in shts:
            if sht.name.upper() == 'POWER RAIL':
                power_sht = sht
            else:
                common_sht = sht

        # 检查是否有Power Rail的sheet
        if power_sht:
            self.power_rail_content_list = power_sht.range('A1').options(expand='table').value
            self.power_rail_list = [x[0] for x in self.power_rail_content_list[1:]]
            self.power_common_power_dict = {xy[0].upper(): xy[1].upper()
                                            for xy in self.power_rail_content_list[1:] if xy[1]}
        else:
            error_message = 'The sheet "Power Rail" is not exists'
            create_error_message(self.error_excel_path, error_message)

        if common_sht is None:
            error_message = 'The sheet "GPIO Sheet" is not exists'
            create_error_message(self.error_excel_path, error_message)

        row = None

        rng = common_sht.range('A2').expand('down').value
        for idx in range(len(rng)):
            try:
                if rng[idx].upper().find('GROUP') > -1:
                    row = idx + 1
                    break
            except AttributeError:
                continue

        # 如果没有GROUP，则报错
        if row is None:
            self.wb.close()
            self.app.quit()
            error_message = 'There is no "Group" split in the table.'
            create_error_message(self.error_excel_path, error_message)
        # 根据row来获取row之后的excel的值
        title_1 = len(common_sht.range((row, 1)).expand('right'))
        title_2 = len(common_sht.range((row + 1, 1)).expand('right'))
        if title_1 > title_2:
            self.common_design_list = common_sht.range((row + 1, 1)).options(expand='table').value
            # 对common design的数据进行处理
            for i in range(len(self.common_design_list)):
                x = self.common_design_list[i]
                for j in range(len(x)):
                    # 跳过值为None的数据
                    if x[j] is not None:
                        # 如果strip后为空字符串也跳过，如果不为空则upper
                        if isinstance(x[j], float):
                            try:
                                x[j] = str(int(x[j]))
                            except:
                                x[j] = str(x[j])
                        if x[j].strip():
                            self.common_design_list[i][j] = self.common_design_list[i][j].upper().strip()
                        else:
                            self.common_design_list[i][j] = self.common_design_list[i][j].strip()
        else:
            self.common_design_list = common_sht.range((row + 1, 1)).options(expand='table').value
            # 对common design的数据进行处理
            for i in range(len(self.common_design_list)):
                x = self.common_design_list[i]
                for j in range(len(x)):
                    # 跳过值为None的数据
                    if x[j] is not None:
                        # 如果strip后为空字符串也跳过，如果不为空则upper
                        if isinstance(x[j], float):
                            try:
                                x[j] = str(int(x[j]))
                            except:
                                x[j] = str(x[j])
                        if x[j].strip():
                            self.common_design_list[i][j] = self.common_design_list[i][j].upper().strip()
                        else:
                            self.common_design_list[i][j] = self.common_design_list[i][j].strip()

        # 取出common design的数据
        self.col_len = len(self.common_design_list[0])
        for x_idx in range(len(self.common_design_list)):
            item_list = self.common_design_list[x_idx]
            pin = item_list[0].upper()
            # 问题：net可能为None吗, 暂定不会为None
            net = item_list[1].upper() if item_list[1] else None
            net = 'RESERVED' if net and net.find('RESERVED') > -1 else net
            net = 'AIO' if net and net.find('AIO') > -1 else net
            pu_pd = item_list[2].upper() if item_list[2] else None
            resistance = str(item_list[3]).upper() if item_list[3] else None
            try:
                voltage = str(int(item_list[4])).upper() if item_list[4] else None
            except:
                voltage = str(item_list[4]).upper() if item_list[4] else None

            # 如果不是分割字母则存入
            if pin.find('GROUP') == -1:
                self.common_design_pin_list.append(pin)
                self.common_design_pin_net_dict[pin] = net
                self.common_design_pin_pu_pd_dict[pin] = pu_pd
                self.common_design_pin_resistance_dict[pin] = resistance
                self.common_design_pin_voltage_dict[pin] = voltage
                # 如果net中包含 AIO 则可直接忽略这行管控
                if net and net.find('AIO') > -1:
                    self.aio_ignore_idx_list.append(x_idx)
            else:
                self.common_design_pin_list.append(pin)

        self.wb.close()
        self.app.quit()

    def get_data_from_version_excel(self, bom_idx):
        """从输入excel中获取所需数据"""
        # 前一版本表格数据
        self.previous_design_pin_net_dict = {}
        self.previous_design_pin_pu_pd_dict = {}
        self.previous_design_pin_resistance_dict = {}
        self.previous_design_pin_voltage_dict = {}
        self.first_line_version_list = []

        # common desgin compare的数据
        self.origin_common_compare_excel_list = []
        self.last_common_compare_list = []
        self.first_two_line_common_compare_list = []
        self.origin_excel_list = []

        if self.gpio_version_input_excel_path is None:
            error_message = 'Input_GPIO_Table.xlsx is not exists.'
            create_error_message(self.error_excel_path, error_message)

        self.app = xw.App(visible=False, add_book=False)
        self.wb = self.app.books.open(self.gpio_version_input_excel_path)
        shts = self.wb.sheets
        common_compare_sht = None
        self.common_sht_list = []
        for sht in shts:
            if sht.name.upper() == 'COMMON COMPARE':
                common_compare_sht = sht
            else:
                self.common_sht_list.append(sht)

        # 如果有common compare说明
        if common_compare_sht:
            self.origin_common_compare_excel_list = common_compare_sht.range('A3').options(expand='table').value
            self.last_common_compare_list = [x[-5:] for x in self.origin_common_compare_excel_list]
            self.first_two_line_common_compare_list = [[x.value for x in xlrd.open_workbook(self.gpio_version_input_excel_path)
                                                           .sheet_by_name('Common Compare').row(0)]]
            self.first_two_line_common_compare_list.append([x.value for x in xlrd.open_workbook(self.gpio_version_input_excel_path)
                                                           .sheet_by_name('Common Compare').row(1)])
            # 对common design的数据进行处理
            for i in range(len(self.last_common_compare_list)):
                x = self.last_common_compare_list[i]
                for j in range(len(x)):
                    # 跳过值为None的数据
                    if x[j] is not None:
                        # 如果strip后为空字符串也跳过，如果不为空则upper
                        if x[j].strip():
                            self.last_common_compare_list[i][j] = self.last_common_compare_list[i][j].upper().strip()
                        else:
                            self.last_common_compare_list[i][j] = self.last_common_compare_list[i][j].strip()
        else:
            error_message = 'The sheet common compare is not exists.'
            create_error_message(self.error_excel_path, error_message)

        # 如果是后续版本，则取出前一版本的表格数据
        if self.common_sht_list and len(self.common_sht_list) == 3:
            self.first_line_version_list = [x.value for x in xlrd.open_workbook
                (self.gpio_version_input_excel_path).sheet_by_name('BOM1').row(0)]
            self.origin_excel_list = self.common_sht_list[bom_idx].range('A2').options(expand='table').value[1:]
            self.col_len = len(self.origin_excel_list[0])
            for x_idx in range(len(self.origin_excel_list)):
                item_list = self.origin_excel_list[x_idx]
                pin = item_list[0].upper()
                net = item_list[self.col_len - 5] if item_list[self.col_len - 5] else None
                pu_pd = item_list[self.col_len - 4] if item_list[self.col_len - 4] else None
                resistance = str(item_list[self.col_len - 3]) if item_list[self.col_len - 3] else None
                voltage = item_list[self.col_len - 2] if item_list[self.col_len - 2] else None

                # 如果不是分割字母则存入
                if pin.find('GROUP') == -1:
                    self.previous_design_pin_net_dict[pin] = net
                    self.previous_design_pin_pu_pd_dict[pin] = pu_pd
                    self.previous_design_pin_resistance_dict[pin] = resistance
                    self.previous_design_pin_voltage_dict[pin] = voltage
        else:
            error_message = 'The sheet BOM1, BOM2, BOM3 is not exists.'
            create_error_message(self.error_excel_path, error_message)
        self.wb.close()
        self.app.quit()


# 读取 pstxnet.dat，pstxprt.dat，pstchip.dat 及 EXP文件的模块
class ExtractIOData:
    """对DSN导出的报告进行数据处理"""

    def __init__(self, root_path, error_excel_path):
        self.root_path = root_path
        self.error_excel_path = error_excel_path

        # pstxnet.dat
        self.all_net_list_ = []
        self.net_node_dict_ = {}
        self.net_node_list_ = []

        # pstxprt.dat
        self.all_node_list_ = []
        self.node_page_dict_ = {}
        self.all_res_dict_ = {}
        self.all_res_list_ = []
        self.all_diode_list_ = []
        self.ic_ext_icname_dict_ = {}

        # pstchip.dat
        self.pin_name_list_ = []
        self.ext_icname_pin_num_dict_ = {}

        # exp
        self.ic_ni_dict_ = {}
        self.ic_value_dict = {}

    def extract_pstxnet(self):
        """提取pstxnet.dat的数据"""
        try:
            with open(os.path.join(self.root_path, 'pstxnet.dat'), 'r') as file1:
                content1 = file1.read().split('NET_NAME')
                for ind1 in range(len(content1)):
                    content1[ind1] = content1[ind1].split('\n')
                for x in content1:
                    node_list = []
                    self.all_net_list_.append(x[1][1:-1])
                    for y_idx in range(len(x)):
                        if x[y_idx].find('NODE_NAME') > -1:
                            node_list.append(
                                [x[y_idx].split('NODE_NAME\t')[-1].split(' ')[0], x[y_idx + 2].split("'")[1]])
                    node_flatten_list = list(flatten([[x[1][1:-1]] + node_list]))
                    # print('node_flatten_list', node_flatten_list)
                    self.net_node_dict_[x[1][1:-1]] = node_flatten_list
                    self.net_node_list_.append(node_flatten_list)
                self.all_net_list_ = self.all_net_list_[1:]

            return self.all_net_list_, self.net_node_list_, self.net_node_dict_
        except FileNotFoundError:
            error_message = 'Missing pstxnet.dat file'
            create_error_message(self.error_excel_path, error_message)

    def extract_pstxprt(self):
        """提取pstxprt.dat的数据"""
        try:
            with open(os.path.join(self.root_path, 'pstxprt.dat'), 'r') as file2:
                content2 = file2.read().split('PART_NAME')

                primitive_list = []

                for ind2 in range(len(content2)):
                    content2[ind2] = content2[ind2].split('\n')

                for x in content2:
                    # print(x)
                    # print('\n')
                    node = x[1].split(' ')[1]
                    self.all_node_list_.append(node)

                    pattern = re.compile(r".*?_.*?_(.*?)_.*?")
                    res_val = pattern.findall(x[1].split(' ')[2])

                    if x[0] == '':
                        pattern2 = re.compile(r".*?'(.*?)'.*?")
                        pattern3 = re.compile(r".*?@.*?@(.*?)\..*?")
                        node1 = x[1].split(' ')[1]
                        self.ic_ext_icname_dict_[node1] = pattern2.findall(x[1])[0]

                        if pattern3.findall(x[5])[0].upper().find('RESISTOR') > -1:
                            self.all_res_list_.append(node1)
                        elif pattern3.findall(x[5])[0].upper().find('DIODE') > -1:
                            self.all_diode_list_.append(node1)

                    if res_val:
                        self.all_res_dict_[node] = res_val[0]

                    if x[0] == '':
                        primitive_list.append(x[1].split("\'")[1])
                    if 'page' in x[6]:
                        page_now = x[6].split(':')[-1].split('_')[0]
                    else:
                        page_now = x[7].split(':')[-1].split('_')[0]
                    if self.node_page_dict_.get(page_now):
                        self.node_page_dict_[page_now] += [self.all_node_list_[-1]]
                    else:
                        self.node_page_dict_[page_now] = [self.all_node_list_[-1]]

            return self.all_node_list_, self.node_page_dict_, \
                   self.all_res_list_, self.all_diode_list_, self.ic_ext_icname_dict_
        except FileNotFoundError:
            error_message = 'Missing pstxprt.dat file'
            create_error_message(self.error_excel_path, error_message)

    def extract_pstchip(self, GPIO_pin_name_list_org=None, func=None):
        try:
            with open(os.path.join(self.root_path, 'pstchip.dat'), 'r') as file5:

                # 用于测试阶段， 对表中的pin name进行变形
                if func == 'reshape_pin_name':
                    content5 = file5.read().split('primitive')
                    pattern1 = re.compile(r".*?'(.*?)':")
                    self.pin_name_list_ = []
                    for y in GPIO_pin_name_list_org:
                        y_flag = True
                        # 先对输入pin进行处理
                        # 去除前后空格
                        y = y.strip()
                        if y.upper().find('GROUP') > -1:
                            self.pin_name_list_.append(y)
                        else:
                            # 对有无/符号进行分类
                            if y.find('/') > -1:
                                y = '/'.join(map(lambda x: x.strip(), y.split('/')))
                                y = y[:31]
                                # print(y)
                            for x in content5:
                                key_items = pattern1.findall(x)
                                for key_idx in range(len(key_items)):

                                    if key_items[key_idx] == y:
                                        self.pin_name_list_.append(y)
                                        # print(1, y)
                                        y_flag = False
                                        break
                                    elif key_items[key_idx].find(y + '/') > -1:
                                        # print(2, key_items[key_idx])
                                        self.pin_name_list_.append(key_items[key_idx])
                                        y_flag = False
                                        break
                                if y_flag is False:
                                    break
                            # pin名错误
                            if y_flag:
                                self.pin_name_list_.append(None)
                    return self.pin_name_list_

                # 用于第一阶段, 获取pin name
                if func == 'get_pin_name':
                    content5 = file5.read().split('primitive')
                    pattern1 = re.compile(r".*?'(.*?)':")
                    self.pin_name_list_ = []
                    for x in content5:
                        key_items = pattern1.findall(x)
                        for key_idx in range(len(key_items)):
                            self.pin_name_list_.append(key_items[key_idx].upper())
                    return sorted(self.pin_name_list_)

                # 获取ic及其pin的数量信息
                if func == 'get_ic_pin_number':
                    content = file5.read().split('end_primitive')
                    pattern = re.compile(r".*?primitive '(.*?)'")

                    for c_item in content:
                        key_item = pattern.findall(c_item)
                        if key_item:
                            self.ext_icname_pin_num_dict_[key_item[0]] = c_item.count('PIN_NUMBER')

                    return self.ext_icname_pin_num_dict_

                # 获取pin name对应的pin location
                if func == 'get_pin_name_pin_location':
                    content5 = file5.read().split('primitive')
                    pattern1 = re.compile(r".*?'(.*?)':")
                    key_list = []
                    value_list = []
                    pattern2 = re.compile(r".*?PIN_NUMBER='\((.*?)\)';")
                    for x in content5:
                        pattern1_data = pattern1.findall(x)
                        if pattern1_data:
                            key_list.append(pattern1_data)
                        pattern2_data = pattern2.findall(x)
                        if pattern2_data:
                            value_half0_list = []
                            for pattern2_data_item in pattern2_data:

                                # 为了处理BA45,0,0,0,0,0,0,0,0,0,0,0,0这种情况
                                if pattern2_data_item.find(',') > -1:
                                    value_half1_list = list(set(pattern2_data_item.split(',')))
                                    try:
                                        value_half1_list.remove('0')
                                        value_half0_list.append(value_half1_list[0])
                                    # 二极管，三极管也有逗号，所以要区分开
                                    except:
                                        value_half0_list.append(pattern2_data_item)

                                else:
                                    value_half0_list.append(pattern2_data_item)
                            value_list.append(value_half0_list)

                    pin_name_pin_location_dict = dict(zip(flatten(key_list), flatten(value_list)))
                    return pin_name_pin_location_dict
        except FileNotFoundError:
            error_message = 'Missing pstchip.dat file'
            create_error_message(self.error_excel_path, error_message)

    def extract_exp(self, file_name):
        with open(os.path.join(self.root_path, file_name), 'r', encoding='gb18030') as file7:
            file7.readline()
            topic_list = file7.readline().split('\t')
            # ADD:添加异常处理机制
            bom_idx = topic_list.index('"BOM"')

            # try:
            for line in file7.readlines():
                line_list = line.split('\t')
                line_id = line_list[1][1:-1]
                line_value = line_list[3][1:-1]

                if '_' in line_value:
                    self.ic_value_dict[line_id] = line_value.split('_')[0]
                else:
                    self.ic_value_dict[line_id] = line_value
                # print(line_id)
                line_ni = line_list[bom_idx][1:-1]
                self.ic_ni_dict_[line_id.upper()] = line_ni
        # print(self.ic_ni_dict_)
        return self.ic_ni_dict_, self.ic_value_dict


# 跑出pin的详细走线信息
class ExtractPinData:

    def __init__(self, sheet):

        self.sheet = sheet

        # 列项信息
        self.pin_name_coord = None
        self.pin_location_coord = None

        # 列项详细信息
        self.GPIO_pin_name_list = []
        self.GPIO_pin_location_list = []

        # 详细走线信息
        self.pin_net_node_list = []
        self.pin_net_node_dict = {}
        self.error_all_list = []

    def get_headline_detail_info(self, pin_name_coord=None, pin_location_coord=None):
        """获取列项的详细信息"""
        # 取得每一个需要check的列的数据并存成list的形式
        group_flag = True
        # 获取要check的sheet中的列项信息
        while group_flag:
            col_idx = len(self.sheet.range(pin_name_coord).options(expand='table').value) + pin_name_coord[0] - 1
            pin_name_list = [str(x) for x in self.sheet.range((pin_name_coord[0], pin_name_coord[1]),
                                                              (col_idx, pin_name_coord[1])).value]

            # pin_location_list = [str(x) for x in self.sheet.range((pin_name_coord[0], pin_location_coord[1]),
            #                                                       (col_idx, pin_location_coord[1])).value]

            self.GPIO_pin_name_list.append(pin_name_list + ['Group'])
            # self.GPIO_pin_location_list.append(pin_location_list + ['Group'])

            # col_idx += 1
            # # 不为none说明下面还有group
            # if self.sheet.range(col_idx + 1, pin_name_coord[1]).value:
            #     pin_name_coord = (col_idx + 1, pin_name_coord[1])
            #     pin_location_coord = (col_idx + 1, pin_location_coord[1])
            # else:
            #     group_flag = False
        return self.GPIO_pin_name_list  # self.GPIO_pin_location_list

    def get_detail_layout_info(self, net_node_list, GPIO_pin_name_list_org,
                               all_res_list, all_diode_list, IC_pin_num_dict, Exclude_Net_List):
        """获取详细走线信息"""
        net_node_copy_list = copy.deepcopy(net_node_list)
        # 对每一个pin name进行数据处理，找出与之对应的走线规则
        for pin_idx in range(len(GPIO_pin_name_list_org)):
            # 遍历pin name去除分割行
            if str(GPIO_pin_name_list_org[pin_idx]).upper().find('GROUP') == -1:
                # 遍历没有拼写错误的所有pin脚
                if GPIO_pin_name_list_org[pin_idx]:
                    node_item_flag = False
                    # final_flag = False
                    no_nc_flag = False
                    for node_item in net_node_list:
                        net_item = node_item[0]
                        # 找到pin脚连接的信号线信息
                        if GPIO_pin_name_list_org[pin_idx] in node_item[1:]:
                            pin_net_node_list1 = [net_item]
                            node_item1 = copy.deepcopy(node_item)
                            node_item1.pop(node_item1.index(GPIO_pin_name_list_org[pin_idx], 1) - 1)
                            node_item1.pop(node_item1.index(GPIO_pin_name_list_org[pin_idx], 1))
                            node_pin_location_item = node_item1[2::2]
                            node_item1 = node_item1[1::2]

                            flagfour = True
                            split_flag = True
                            split_node_list = []
                            split_pin_location_list = []
                            layer_num = 0
                            layer_add_num_dict = {}

                            # 如果匹配到NC，因为NC在最后，所以匹配到NC说明前面都不匹配
                            if pin_net_node_list1[0] == 'NC':
                                # 如果只匹配到NC
                                if no_nc_flag is False:
                                    # if str(GPIO_pupd_list_org[pin_idx]) != 'None':
                                    pin_net_node_list1 = []
                                    # print('NC', pin_net_node_list1)
                                    self.pin_net_node_dict[GPIO_pin_name_list_org[pin_idx]] = pin_net_node_list1
                                    node_item_flag = True
                                # 如果除了NC还匹配到其他
                                # 这段代码好像没用，但是我也不想动它了
                                # else:
                                #     final_flag = True
                            else:
                                no_nc_flag = True
                                # final_flag = False
                                while flagfour:
                                    node_item3 = []

                                    break_flag = False
                                    split_out_flag = False
                                    # if pin_name_list[pin_idx] == 'GPIO1':
                                    #     print(1, node_item1)
                                    if split_flag:
                                        split_node_list.append(copy.deepcopy(node_item1))
                                        split_pin_location_list.append(copy.deepcopy(node_pin_location_item))
                                    for x_idx in range(len(node_item1)):
                                        # IC_flag = False
                                        next_flag = False
                                        all_break = False
                                        add_num = 0
                                        layer_num += 1

                                        item1 = node_item1[x_idx]
                                        # print(2, item1)
                                        split_node_list[-1].pop(split_node_list[-1].index(item1))
                                        split_pin_location_list[-1].pop(split_pin_location_list[-1].
                                                                        index(node_pin_location_item[x_idx]))
                                        # 对下一个经过的元器件是否是NI进行判断
                                        # # if pin_idx not in consider_ni_idx_list:
                                        # if len(ic_ni_dict[item1].split(',')) == 3:
                                        #     ic_ni = ic_ni_dict[item1].split(',')[idx]
                                        # else:
                                        #     ic_ni = ic_ni_dict[item1]
                                        # if ic_ni == 'NI':
                                        #     pin_net_node_list1.append('NI')
                                        #     split_node_list.append([])
                                        #     split_pin_location_list.append([])
                                        #     split_flag = False
                                        #     add_num += 1
                                        # # 如果上电
                                        # else:
                                        # 大于4并且不是排阻则说明到另外一个芯片了，停止
                                        if item1 not in all_res_list and IC_pin_num_dict[item1] >= 3:
                                            split_flag = False
                                            pin_net_node_list1.append(item1)
                                            add_num += 1
                                            split_node_list.append([])
                                            split_pin_location_list.append([])
                                            # 如果下一个芯片是二极管，要考虑二极管的单向导通性，由A到K
                                            # 如果为K就不朝下走了
                                        elif item1 in all_diode_list and \
                                                node_pin_location_item[x_idx].upper() == 'K':
                                            split_flag = False
                                            pin_net_node_list1.append(item1)
                                            add_num += 1
                                            split_node_list.append([])
                                            split_pin_location_list.append([])
                                        else:
                                            #
                                            # print(pin_net_node_list1)
                                            # 判断是否为终止端元器件（中间的元器件会出现两次）
                                            if node_item1.count(item1) >= 1:
                                                # count = 0
                                                for node_item2 in net_node_copy_list:
                                                    # count += 1
                                                    add_sch_flag = False
                                                    # 找到元器件所连接的另一根线
                                                    # 如果这次经过的线与上次或第一次相同，则退出
                                                    if item1 in node_item2 and node_item2 != node_item \
                                                            and node_item2 != node_item3:

                                                        # 如果中间没有经过过这个元器件则进入循环
                                                        if node_item2[0] not in pin_net_node_list1:
                                                            add_sch_flag = True
                                                            pin_net_node_list1.append(item1)
                                                            pin_net_node_list1.append(node_item2[0])
                                                            add_num += 2
                                                            if node_item2[0] != node_item[0] and node_item2 \
                                                                    != node_item3:
                                                                if node_item2[0] in Exclude_Net_List:
                                                                    split_node_list.append([])
                                                                    split_pin_location_list.append([])
                                                                    split_flag = False
                                                                    break

                                                                node_item1 = copy.deepcopy(node_item2)
                                                                node_item3 = copy.deepcopy(node_item2)

                                                                node_item1.pop(node_item1.index(item1) - 1)
                                                                node_item1.pop(node_item1.index(item1))

                                                                node_pin_location_item = node_item1[2::2]
                                                                node_item1 = node_item1[1::2]

                                                                next_flag = True
                                                                split_flag = True
                                                                break_flag = True
                                                                break
                                                                # break 不要break，是因为可能元器件有超过两个pin，
                                                                # 要所有都遍历到, 虽然速度会变慢
                                                            else:
                                                                all_break = True

                                                    if net_node_copy_list[
                                                        -1] == node_item2 and add_sch_flag is \
                                                            False:
                                                        split_flag = False
                                                        add_num += 1
                                                        split_node_list.append([])
                                                        split_pin_location_list.append([])
                                                        pin_net_node_list1.append(item1)

                                                        if pin_net_node_list1 not in self.pin_net_node_list:
                                                            self.pin_net_node_list.append(pin_net_node_list1)

                                                        if self.pin_net_node_dict.get(
                                                                GPIO_pin_name_list_org[pin_idx]):
                                                            pin_net_dict_list = self.pin_net_node_dict[
                                                                GPIO_pin_name_list_org[pin_idx]]
                                                            pin_net_dict_list.append(pin_net_node_list1)
                                                            self.pin_net_node_dict[GPIO_pin_name_list_org[
                                                                pin_idx]] = pin_net_dict_list
                                                        else:
                                                            self.pin_net_node_dict[GPIO_pin_name_list_org[
                                                                pin_idx]] = [pin_net_node_list1]

                                        if all_break:
                                            pass
                                        else:
                                            layer_add_num_dict[layer_num] = add_num
                                            split_node_flag = True
                                            before_layer_num = 0
                                            if next_flag is False:
                                                if split_flag is False or node_item1[-1] == item1:
                                                    split_flag = True
                                                    split_node_list.pop(-1)
                                                    split_pin_location_list.pop(-1)
                                                    before_layer_num = copy.deepcopy(layer_num)
                                                    layer_num = len(split_node_list)
                                                    if split_node_list:
                                                        try:
                                                            while not split_node_list[-1]:
                                                                split_node_list.pop(-1)
                                                                split_pin_location_list.pop(-1)
                                                                layer_num -= 1
                                                                # print('layer_num', layer_num)
                                                        except IndexError:
                                                            pass

                                                    if split_node_list:
                                                        node_item1 = split_node_list[-1]
                                                        node_pin_location_item = split_pin_location_list[-1]
                                                        # if len(node_item1) == 1:
                                                        split_node_list.pop(-1)
                                                        split_pin_location_list.pop(-1)
                                                        # if node_item1[-1] == item1:
                                                        break_flag = True
                                                        split_node_flag = True

                                                        # print('split_node_flag', split_node_flag)
                                                    else:
                                                        split_node_flag = False
                                                        flagfour = False
                                                        if node_item1[-1] == item1:
                                                            split_out_flag = True

                                                    if pin_net_node_list1 not in self.pin_net_node_list:
                                                        self.pin_net_node_list.append(pin_net_node_list1)

                                                    if 'NI' in pin_net_node_list1:
                                                        if self.pin_net_node_dict.get(GPIO_pin_name_list_org[pin_idx]):
                                                            pass
                                                        else:
                                                            self.pin_net_node_dict[
                                                                GPIO_pin_name_list_org[pin_idx]] = []
                                                    else:
                                                        if self.pin_net_node_dict.get(GPIO_pin_name_list_org[pin_idx]):
                                                            pin_net_dict_list = self.pin_net_node_dict[
                                                                GPIO_pin_name_list_org[pin_idx]]
                                                            pin_net_dict_list.append(pin_net_node_list1)
                                                            self.pin_net_node_dict[GPIO_pin_name_list_org[pin_idx]] = \
                                                                pin_net_dict_list
                                                        else:
                                                            self.pin_net_node_dict[GPIO_pin_name_list_org[pin_idx]] = \
                                                                [pin_net_node_list1]

                                            if split_node_flag:
                                                if before_layer_num:
                                                    for layer_idx in range(layer_num, before_layer_num + 1):
                                                        # print(layer_idx)
                                                        if layer_add_num_dict[layer_idx] != 0:
                                                            pin_net_node_list1 = pin_net_node_list1[:
                                                                                                    -
                                                                                                    layer_add_num_dict[
                                                                                                        layer_idx]]

                                                        layer_add_num_dict.pop(layer_idx)
                                                    layer_num -= 1
                                            if break_flag:
                                                break

                                            if split_out_flag:
                                                break

                            # if node_item == net_node_list[-1] and final_flag:
                            #     node_item_flag = True
                            if node_item_flag:
                                break

                else:
                    self.error_all_list.append(pin_idx)
                    self.pin_net_node_dict[GPIO_pin_name_list_org[pin_idx]] = [['misspelled']]
        return self.pin_net_node_list, self.pin_net_node_dict, self.error_all_list


# 生成完整版 GPIO TABLE
def generate_report():
    """生成报告"""
    # 从input excel中读取数据
    input_handler = InputExcelHandler()
    input_handler.get_input_excel_path()
    input_path = input_handler._root_path
    gpio_version_input_excel_path = input_handler.gpio_version_input_excel_path
    error_excel_path = input_handler.error_excel_path
    output_path = input_handler._output_path

    # 设置初版flag
    first_version_flag = True
    bom_idx_bom_content_dict = {}
    # print(common_compare_sht)
    # 如果 gpio_version_input_excel 为None，说明是初版
    input_handler.get_data_from_common_design_excel()
    # 获取common design表格数据
    common_design_pin_list = input_handler.common_design_pin_list
    common_design_pin_net_dict = input_handler.common_design_pin_net_dict
    common_design_pin_pu_pd_dict = input_handler.common_design_pin_pu_pd_dict
    common_design_pin_resistance_dict = input_handler.common_design_pin_resistance_dict
    common_design_pin_voltage_dict = input_handler.common_design_pin_voltage_dict
    common_design_list = input_handler.common_design_list
    origin_excel_list = common_design_list
    # 获取power rail表格数据
    power_rail_list = input_handler.power_rail_list
    aio_ignore_idx_list = input_handler.aio_ignore_idx_list
    power_common_power_dict = input_handler.power_common_power_dict

    col_len = input_handler.col_len
    # 如果有 gpio_version_input_excel 说明为后续版本
    if gpio_version_input_excel_path is not None:
        first_version_flag = False

        # BOM1 数据
        input_handler.get_data_from_version_excel(0)
        previous_design_pin_net_dict_bom1 = input_handler.previous_design_pin_net_dict
        previous_design_pin_pu_pd_dict_bom1 = input_handler.previous_design_pin_pu_pd_dict
        previous_design_pin_resistance_dict_bom1 = input_handler.previous_design_pin_resistance_dict
        previous_design_pin_voltage_dict_bom1 = input_handler.previous_design_pin_voltage_dict
        origin_excel_list_bom1 = input_handler.origin_excel_list

        bom_idx_bom_content_dict[1] = [origin_excel_list_bom1, previous_design_pin_net_dict_bom1,
                                       previous_design_pin_pu_pd_dict_bom1, previous_design_pin_resistance_dict_bom1,
                                       previous_design_pin_voltage_dict_bom1]

        # BOM2 数据
        input_handler.get_data_from_version_excel(1)
        previous_design_pin_net_dict_bom2 = input_handler.previous_design_pin_net_dict
        previous_design_pin_pu_pd_dict_bom2 = input_handler.previous_design_pin_pu_pd_dict
        previous_design_pin_resistance_dict_bom2 = input_handler.previous_design_pin_resistance_dict
        previous_design_pin_voltage_dict_bom2 = input_handler.previous_design_pin_voltage_dict
        origin_excel_list_bom2 = input_handler.origin_excel_list

        bom_idx_bom_content_dict[2] = [origin_excel_list_bom2, previous_design_pin_net_dict_bom2,
                                       previous_design_pin_pu_pd_dict_bom2, previous_design_pin_resistance_dict_bom2,
                                       previous_design_pin_voltage_dict_bom2]

        # BOM3 数据
        input_handler.get_data_from_version_excel(2)
        previous_design_pin_net_dict_bom3 = input_handler.previous_design_pin_net_dict
        previous_design_pin_pu_pd_dict_bom3 = input_handler.previous_design_pin_pu_pd_dict
        previous_design_pin_resistance_dict_bom3 = input_handler.previous_design_pin_resistance_dict
        previous_design_pin_voltage_dict_bom3 = input_handler.previous_design_pin_voltage_dict
        origin_excel_list_bom3 = input_handler.origin_excel_list

        bom_idx_bom_content_dict[3] = [origin_excel_list_bom3, previous_design_pin_net_dict_bom3,
                                       previous_design_pin_pu_pd_dict_bom3, previous_design_pin_resistance_dict_bom3,
                                       previous_design_pin_voltage_dict_bom3]

        # 获取common compare数据
        origin_common_compare_excel_list = input_handler.origin_common_compare_excel_list
        last_common_compare_list = input_handler.last_common_compare_list
        first_two_line_common_compare_list = input_handler.first_two_line_common_compare_list
        first_line_version_list = input_handler.first_line_version_list

        col_len = input_handler.col_len

    # 输入表格的通用数据

    GND_Net_List = ['GND', 'AGND']
    all_power_rail_list = power_rail_list + GND_Net_List

    group_idx_list = [idx for idx in range(len(common_design_pin_list))
                      if common_design_pin_list[idx].upper().find('GROUP') > -1]
    # 生成读取数据的类
    extractIOData = ExtractIOData(input_path, error_excel_path)

    # 生成结果表格的绝对路径
    output_excel_path = os.path.join(output_path, 'Output_GPIO_Table.xlsx')

    all_pin_list = extractIOData.extract_pstchip(func='get_pin_name')
    # 首先对pin name进行判断，确定输入excel汇总所有的pin名称都正确
    error_pin_name_list = [x for x in common_design_pin_list if x not in all_pin_list and x.find('GROUP') == -1]
    # print('error_pin_name_list', error_pin_name_list)
    # 如果存在错误则生成错误报告excel
    if error_pin_name_list:
        create_error_message(error_excel_path, error_pin_name_list)

    all_net_list, net_node_list, net_node_dict = extractIOData.extract_pstxnet()
    all_node_list, node_page_dict, all_res_list, all_diode_list, ic_ext_icname_dict = \
        extractIOData.extract_pstxprt()
    extractPinData = ExtractPinData(None)
    # 输出每个IC的pin脚数目
    ext_icname_pin_num_dict = extractIOData.extract_pstchip(func='get_ic_pin_number')
    # Exclude_Net_List, PWR_Net_List, GND_Net_List = get_exclude_netlist(all_net_list)
    # 寻找.exp文件
    file_name = ''
    for x in os.listdir(input_path):
        (shotname, extension) = os.path.splitext(x)
        # print(extension)
        if extension == '.EXP':
            file_name = x
            file_base_name = os.path.splitext(x)[0]

    # 如果没有.exp文件，抛出异常
    if file_name == '':
        error_message = 'Missing *.EXP file'
        create_error_message(error_excel_path, error_message)

    # 生成ic BOM列表
    ic_ni_dict, ic_value_dict = extractIOData.extract_exp(file_name)
    # 获得所有电源线
    IC_pin_num_dict = {}
    for ic_item in ic_ext_icname_dict.keys():
        IC_pin_num_dict[ic_item] = ext_icname_pin_num_dict[ic_ext_icname_dict.get(ic_item)]

    # 分三种情况
    # 将生成的现实数据输出做比较
    # 創建excel
    workbook = xlsxwriter.Workbook(output_excel_path)
    first_title = ['Pin name']
    title_list = ['Signal name', 'PU/PD', 'Resistance', 'Power Rail', 'Note']

    # 标题格式
    title_format = workbook.add_format({'font_size': 12, 'bold': True, 'bg_color': '#9BC2E6', 'border': 1})
    column_group_format = workbook.add_format({'font_size': 12, 'bg_color': '#FFE699', 'border': 1})
    # 版本号格式
    version_format = workbook.add_format({'align': 'center', 'valign': 'vcenter',
                                          'bg_color': '#F99FCC', 'font_size': 22, 'bold': True, 'border': 1})
    # 内容格式
    column_pin_name_format = workbook.add_format({'font_size': 12, 'bg_color': '#65DADD', 'border': 1})
    column_note_format = workbook.add_format({'font_size': 12, 'border': 1})
    column_pass_format = workbook.add_format({'font_size': 12, 'bg_color': '#92D050', 'border': 1})
    column_check_format = workbook.add_format({'font_size': 12, 'bg_color': '#FFFF00', 'border': 1})
    column_fail_format = workbook.add_format({'font_size': 12, 'bg_color': '#FC2443', 'border': 1})

    # 列项信息列表
    column_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
                   'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                   'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ',
                   'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
                   'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ',
                   'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ',
                   'CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ',
                   'CR', 'CS', 'CT', 'CU', 'CV', 'CW', 'CX', 'CY', 'CZ',
                   'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN', 'DO', 'DP', 'DQ',
                   'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX', 'DY', 'DZ',
                   'EA', 'EB', 'EC', 'ED', 'EE', 'EF', 'EG', 'EH', 'EI', 'EJ', 'EK', 'EL', 'EM', 'EN', 'EO', 'EP', 'EQ',
                   'ER', 'ES', 'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ',
                   'FA', 'FB', 'FC', 'FD', 'FE', 'FF', 'FG', 'FH', 'FI', 'FJ', 'FK', 'FL', 'FM', 'FN', 'FO', 'FP', 'FQ',
                   'FR', 'FS', 'FT', 'FU', 'FV', 'FW', 'FX', 'FY', 'FZ',
                   'GA', 'GB', 'GC', 'GD', 'GE', 'GF', 'GG', 'GH', 'GI', 'GJ', 'GK', 'GL', 'GM', 'GN', 'GO', 'GP', 'GQ',
                   'GR', 'GS', 'GT', 'GU', 'GV', 'GW', 'GX', 'GY', 'GZ',
                   'HA', 'HB', 'HC', 'HD', 'HE', 'HF', 'HG', 'HH', 'HI', 'HJ', 'HK', 'HL', 'HM', 'HN', 'HO', 'HP', 'HQ',
                   'HR', 'HS', 'HT', 'HU', 'HV', 'HW', 'HX', 'HY', 'HZ',
                   'IA', 'IB', 'IC', 'ID', 'IE', 'IF', 'IG', 'IH', 'II', 'IJ', 'IK', 'IL', 'IM', 'IN', 'IO', 'IP', 'IQ',
                   'IR', 'IS', 'IT', 'IU', 'IV', 'IW', 'IX', 'IY', 'IZ',
                   'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI', 'JJ', 'JK', 'JL', 'JM', 'JN', 'JO', 'JP', 'JQ',
                   'JR', 'JS', 'JT', 'JU', 'JV', 'JW', 'JX', 'JY', 'JZ',
                   'KA', 'KB', 'KC', 'KD', 'KE', 'KF', 'KG', 'KH', 'KI', 'KJ', 'KK', 'KL', 'KM', 'KN', 'KO', 'KP', 'KQ',
                   'KR', 'KS', 'KT', 'KU', 'KV', 'KW', 'KX', 'KY', 'KZ',
                   'LA', 'LB', 'LC', 'LD', 'LE', 'LF', 'LG', 'LH', 'LI', 'LJ', 'LK', 'LL', 'LM', 'LN', 'LO', 'LP', 'LQ',
                   'LR', 'LS', 'LT', 'LU', 'LV', 'LW', 'LX', 'LY', 'LZ',
                   'MA', 'MB', 'MC', 'MD', 'ME', 'MF', 'MG', 'MH', 'MI', 'MJ', 'MK', 'ML', 'MM', 'MN', 'MO', 'MP', 'MQ',
                   'MR', 'MS', 'MT', 'MU', 'MV', 'MW', 'MX', 'MY', 'MZ',
                   'NA', 'NB', 'NC', 'ND', 'NE', 'NF', 'NG', 'NH', 'NI', 'NJ', 'NK', 'NL', 'NM', 'NN', 'NO', 'NP', 'NQ',
                   'NR', 'NS', 'NT', 'NU', 'NV', 'NW', 'NX', 'NY', 'NZ',
                   'OA', 'OB', 'OC', 'OD', 'OE', 'OF', 'OG', 'OH', 'OI', 'OJ', 'OK', 'OL', 'OM', 'ON', 'OO', 'OP', 'OQ',
                   'OR', 'OS', 'OT', 'OU', 'OV', 'OW', 'OX', 'OY', 'OZ',
                   'PA', 'PB', 'PC', 'PD', 'PE', 'PF', 'PG', 'PH', 'PI', 'PJ', 'PK', 'PL', 'PM', 'PN', 'PO', 'PP', 'PQ',
                   'PR', 'PS', 'PT', 'PU', 'PV', 'PW', 'PX', 'PY', 'PZ',
                   'QA', 'QB', 'QC', 'QD', 'QE', 'QF', 'QG', 'QH', 'QI', 'QJ', 'QK', 'QL', 'QM', 'QN', 'QO', 'QP', 'QQ',
                   'QR', 'QS', 'QT', 'QU', 'QV', 'QW', 'QX', 'QY', 'QZ',
                   'RA', 'RB', 'RC', 'RD', 'RE', 'RF', 'RG', 'RH', 'RI', 'RJ', 'RK', 'RL', 'RM', 'RN', 'RO', 'RP', 'RQ',
                   'RR', 'RS', 'RT', 'RU', 'RV', 'RW', 'RX', 'RY', 'RZ',
                   'SA', 'SB', 'SC', 'SD', 'SE', 'SF', 'SG', 'SH', 'SI', 'SJ', 'SK', 'SL', 'SM', 'SN', 'SO', 'SP', 'SQ',
                   'SR', 'SS', 'ST', 'SU', 'SV', 'SW', 'SX', 'SY', 'SZ',
                   'TA', 'TB', 'TC', 'TD', 'TE', 'TF', 'TG', 'TH', 'TI', 'TJ', 'TK', 'TL', 'TM', 'TN', 'TO', 'TP', 'TQ',
                   'TR', 'TS', 'TT', 'TU', 'TV', 'TW', 'TX', 'TY', 'TZ',
                   'UA', 'UB', 'UC', 'UD', 'UE', 'UF', 'UG', 'UH', 'UI', 'UJ', 'UK', 'UL', 'UM', 'UN', 'UO', 'UP', 'UQ',
                   'UR', 'US', 'UT', 'UU', 'UV', 'UW', 'UX', 'UY', 'UZ',
                   'VA', 'VB', 'VC', 'VD', 'VE', 'VF', 'VG', 'VH', 'VI', 'VJ', 'VK', 'VL', 'VM', 'VN', 'VO', 'VP', 'VQ',
                   'VR', 'VS', 'VT', 'VU', 'VV', 'VW', 'VX', 'VY', 'VZ',
                   'WA', 'WB', 'WC', 'WD', 'WE', 'WF', 'WG', 'WH', 'WI', 'WJ', 'WK', 'WL', 'WM', 'WN', 'WO', 'WP', 'WQ',
                   'WR', 'WS', 'WT', 'WU', 'WV', 'WW', 'WX', 'WY', 'WZ',
                   'XA', 'XB', 'XC', 'XD', 'XE', 'XF', 'XG', 'XH', 'XI', 'XJ', 'XK', 'XL', 'XM', 'XN', 'XO', 'XP', 'XQ',
                   'XR', 'XS', 'XT', 'XU', 'XV', 'XW', 'XX', 'XY', 'XZ',
                   'YA', 'YB', 'YC', 'YD', 'YE', 'YF', 'YG', 'YH', 'YI', 'YJ', 'YK', 'YL', 'YM', 'YN', 'YO', 'YP', 'YQ',
                   'YR', 'YS', 'YT', 'YU', 'YV', 'YW', 'YX', 'YY', 'YZ',
                   'ZA', 'ZB', 'ZC', 'ZD', 'ZE', 'ZF', 'ZG', 'ZH', 'ZI', 'ZJ', 'ZK', 'ZL', 'ZM', 'ZN', 'ZO', 'ZP', 'ZQ',
                   'ZR', 'ZS', 'ZT', 'ZU', 'ZV', 'ZW', 'ZX', 'ZY', 'ZZ'
                   ]

    # 生成common compare表格
    com_worksheet = workbook.add_worksheet('Common Compare')
    col_idx = (col_len - 1) // 5
    common_compare_diff_flag = False
    if first_version_flag:
        # 如果是第一版，直接生成commom design表格
        first_row = 2
        # 先生成版本号
        com_worksheet.merge_range('B1:E1', 'Common Desgin X00', version_format)
        # 写入标题
        com_worksheet.write_row('A2', first_title + title_list[:-1], title_format)
        # 写入详细数据，如果是group行
        for x in common_design_list:
            x = x[:-1]
            # 如果是group行，填入group颜色
            first_row += 1
            if x[0].strip().upper().find('GROUP') > -1:
                com_worksheet.write_row('A{}'.format(first_row), x, column_group_format)
                continue

            # 如果是pin name列，填入pin name 颜色
            com_worksheet.write('A{}'.format(first_row), x[0], column_pin_name_format)
            com_worksheet.write_row('B{0}:E{0}'.format(first_row), x[1:], column_pass_format)
    else:
        no_group_last_common_compare_list = []
        no_group_common_design_list = []

        common_design_count = len(first_two_line_common_compare_list[0]) // 5
        col_idx = (col_len - 1) // 5

        # 去掉group所在行
        for x in last_common_compare_list:
            if x[0].upper().find('GROUP') == -1:
                no_group_last_common_compare_list.append(x)
        for y in common_design_list:
            y = y[:-1]
            if y[0].upper().find('GROUP') == -1:
                no_group_common_design_list.append(y)

        # 先判断是否此次common design与上一版
        version_num = 0
        first_line_len = len(first_line_version_list)
        for first_idx in range(first_line_len):
            version_item = first_line_version_list[first_idx]
            if version_item:
                if version_item != 'Common Design':
                    version_num += 1
        # 如果与上一版一致
        if no_group_last_common_compare_list == no_group_common_design_list:
            common_compare_diff_flag = False
            # 生成原来的common compare 的版本号
            for x in range(common_design_count - 1):
                com_worksheet.merge_range('{}1:{}1'.format(column_list[x * 5 + 1], column_list[x * 5 + 4]),
                                          first_two_line_common_compare_list[0][x * 5 + 1], version_format)

            # 最后一个版本号要加入最新版本
            com_worksheet.merge_range('{}1:{}1'.format(column_list[(common_design_count - 1) * 5 + 1],
                                                       column_list[(common_design_count - 1) * 5 + 4]),
                                      first_two_line_common_compare_list[0][(common_design_count - 1) * 5 + 1][:17]
                                      + ' ~ X{0:02d}'.format(version_num), version_format)
            # 写入标题
            com_worksheet.write_row('A2', (first_title + title_list[:-1]) * common_design_count, title_format)
        else:
            # 如果与上一版不一致
            common_compare_diff_flag = True
            # 生成原来的common compare 的版本号
            for x in range(common_design_count):
                com_worksheet.merge_range('{}1:{}1'.format(column_list[x * 5 + 1], column_list[x * 5 + 4]),
                                          first_two_line_common_compare_list[0][x * 5 + 1], version_format)
            # 写入标题
            com_worksheet.write_row('A2', (first_title + title_list[:-1]) * (common_design_count + 1), title_format)

        next_row = 2
        # 写入原数据
        column_len = len(origin_common_compare_excel_list[1])
        for x in origin_common_compare_excel_list:
            next_row += 1
            # 如果是group行，填入group颜色
            if x[0].strip().upper().find('GROUP') > -1:
                com_worksheet.write_row('A{}'.format(next_row), x, column_group_format)
                continue

            for y in range(column_len):
                # 如果是pin name列，填入pin name 颜色
                if y % 5 == 0:
                    com_worksheet.write('{}{}'.format(column_list[y], next_row), x[y], column_pin_name_format)
                else:
                    com_worksheet.write('{}{}'.format(column_list[y], next_row), x[y], column_pass_format)

        # 将之前数据错误的变成红色
        open_book = openpyxl.load_workbook(gpio_version_input_excel_path)  # 读取excel
        open_sheet = open_book.get_sheet_by_name('Common Compare')  # 读取Sheet
        rows, cols = open_sheet.max_row, open_sheet.max_column
        for i in range(1, rows):
            for j in range(1, cols):
                ce = open_sheet.cell(row=i, column=j)
                fill = ce.fill
                if fill.start_color.rgb == 'FFFC2443':
                    com_worksheet.write(i - 1, j - 1, ce.value, column_fail_format)

        # 如果有差异，则生成新版本
        if common_compare_diff_flag:
            # 生成新添加的 common compare 的版本号
            com_worksheet.merge_range('{}1:{}1'.format(column_list[common_design_count * 5 + 1],
                                                       column_list[common_design_count * 5 + 4]),
                                      'Common Design X{0:02d}'.format(version_num), version_format)

            # 一个一个cell比较差异
            for x1 in range(len(last_common_compare_list)):
                last_row_list = last_common_compare_list[x1]
                common_design_row_list = common_design_list[x1][:-1]
                # 如果是group行，直接填入，不用比较
                if last_row_list[0].strip().upper().find('GROUP') > -1:
                    com_worksheet.write_row('{}{}'.format(column_list[column_len], x1 + 3),
                                            common_design_row_list, column_group_format)
                    continue
                for y1 in range(len(last_row_list)):
                    # 如果不相等，则标红
                    if last_row_list[y1] != common_design_row_list[y1]:
                        com_worksheet.write('{}{}'.format(column_list[column_len + y1], x1 + 3),
                                            common_design_row_list[y1], column_fail_format)
                    # 如果是 pin name，设置pin name颜色
                    elif y1 == 0:
                        com_worksheet.write('{}{}'.format(column_list[column_len + y1], x1 + 3),
                                            common_design_row_list[y1], column_pin_name_format)
                    else:
                        com_worksheet.write('{}{}'.format(column_list[column_len + y1], x1 + 3),
                                            common_design_row_list[y1], column_pass_format)

    # print(bom_idx_bom_content_dict)
    # 生成三个bom sheet
    for idx in range(3):
        bom_idx = idx + 1
        sheet_name = 'BOM' + str(idx + 1)
        worksheet = workbook.add_worksheet(sheet_name)
        # 获取详细的走线信息
        # 对NI的情况进行判断
        pin_net_node_list, pin_net_node_dict, error_all_list = extractPinData.get_detail_layout_info(
            net_node_list, common_design_pin_list, all_res_list, all_diode_list, IC_pin_num_dict, all_power_rail_list)
        if first_version_flag is False:
            origin_excel_list, previous_design_pin_net_dict_bom, previous_design_pin_pu_pd_dict_bom, \
            previous_design_pin_resistance_dict_bom, previous_design_pin_voltage_dict_bom \
                = bom_idx_bom_content_dict[bom_idx]

        # 对pin_net_node_dict中的value进行去重
        for key_item in pin_net_node_dict.keys():
            pin_net_node_dict[key_item] = list(set([tuple(t) for t in pin_net_node_dict[key_item]]))
        # 通过GPI/O这一项的信息来筛选详细走线信息
        real_signal_name_list = []
        real_resistance_list = []
        real_pu_pd_list = []
        real_power_list = []

        # ************************************************ 生成result及错误信息 ***************************************
        # 生成错误信息
        result_list = []
        error_message_list = []

        # 生成每条pin的表格数据
        for gpp_idx in range(len(common_design_pin_list)):
            pin_name_item = common_design_pin_list[gpp_idx]
            result_flag = True
            pwr_num = 0
            gnd_num = 0
            I_pwr_num = 0
            I_gnd_num = 0
            pwr_item = ''
            gnd_item = ''
            pwr_rail_list = []
            pwr_reserved_list = []
            gnd_rail_list = []
            gnd_reserved_list = []
            pwr_item_list = []
            gnd_item_list = []
            pwr_res_item_list = []
            gnd_res_item_list = []
            half_signal_name_list = []

            # 跑出实际的走线信息
            if pin_name_item.upper().find('GROUP') == -1:
                check_net = common_design_pin_net_dict[pin_name_item]
                check_pu_pd = common_design_pin_pu_pd_dict[pin_name_item]
                check_resistance = common_design_pin_resistance_dict[pin_name_item]
                try:
                    check_resistance = str(int(float(check_resistance)))
                except:
                    pass
                check_voltage = common_design_pin_voltage_dict[pin_name_item]
                net_node_item = pin_net_node_dict[pin_name_item]
                check_common_design_list = [check_net, check_pu_pd, check_resistance, check_voltage]
                # ***************************** 跑出实际数据 *****************************
                # 判断是上拉还是下拉还是分压
                for net_item in net_node_item:
                    # 先判断实际是reserved还是非reserved
                    reserved_flag = False
                    ic_list = list(net_item)[1::2]
                    for ic in ic_list:
                        # 如果有三个NI,I,NI
                        if len(ic_ni_dict.get(ic, 'I').split(',')) == 3:
                            ic_ni = ic_ni_dict.get(ic, 'I').split(',')[idx]
                            if ic_ni.upper() == 'NI':
                                reserved_flag = True
                                break
                        else:
                            if ic_ni_dict.get(ic, 'I') == 'NI':
                                reserved_flag = True
                                break
                    # 如果找到了上拉电源
                    if net_item[-1] in power_rail_list:
                        # 如果power为I并且直接连接电阻
                        if not reserved_flag and net_item[-2] in all_res_list:
                            I_pwr_num += 1
                        pwr_num += 1
                        rail_item = ', ' + net_item[-1]
                        pwr_item += rail_item if pwr_num > 1 else net_item[-1]
                        pwr_rail_list.append(net_item[-1])
                        pwr_reserved_list.append(reserved_flag)
                        pwr_item_list.append(net_item[:-2])
                        if net_item[-2] in all_res_list:
                            pwr_res_item_list.append(ic_value_dict[net_item[-2]])
                        # elif net_item_length > 3 and net_item[-4] in all_res_list:
                        #     print(net_item[0])
                        # pwr_res_item_list.append(ic_value_dict[net_item[-3]])
                        else:
                            pwr_res_item_list.append(None)

                    # 找到了下拉地线
                    if net_item[-1] in GND_Net_List:
                        rail_item = ', ' + net_item[-1]
                        # 如果gnd为I并且直接连接电阻
                        if not reserved_flag and net_item[-2] in all_res_list:
                            I_gnd_num += 1
                        gnd_num += 1
                        gnd_item += rail_item if pwr_num > 1 else net_item[-1]
                        gnd_rail_list.append(net_item[-1])
                        gnd_reserved_list.append(reserved_flag)
                        gnd_item_list.append(net_item[:-2])
                        if net_item[-2] in all_res_list:
                            gnd_res_item_list.append(ic_value_dict[net_item[-2]])
                        # elif net_item_length > 3 and net_item[-4] in all_res_list:
                        #     print(net_item[0])
                        #     gnd_res_item_list.append(ic_value_dict[net_item[-4]])
                        else:
                            gnd_res_item_list.append(None)

                    # 保存出pin的线
                for node_items in net_node_item:
                    if node_items[0] not in half_signal_name_list:
                        half_signal_name_list.append(node_items[0])
                    if len(node_items) >= 3:
                        if node_items[2] not in all_power_rail_list + half_signal_name_list:
                            half_signal_name_list.append(node_items[2])

                if half_signal_name_list:
                    real_signal_name_list.append(half_signal_name_list)
                else:
                    # real_resistance_list.append(None)
                    # real_pu_pd_list.append(None)
                    # real_power_list.append(None)
                    # error_message_list.append(None)
                    real_signal_name_list.append(['NC'])
                # 如果BOM为I的两个PU或者两个PD的情况，不去判断，直接fail
                if I_pwr_num >= 2 or I_gnd_num >= 2:
                    real_resistance_str_list = []
                    real_pu_pd_str_list = []
                    real_power_str_list = []
                    # 如果有pu
                    if pwr_rail_list:
                        for pwr_idx in range(len(pwr_rail_list)):
                            # 判断是否有电阻
                            if pwr_res_item_list[pwr_idx]:
                                # 如果有电阻则存入，否则不存入
                                real_resistance_str_list.append(pwr_res_item_list[pwr_idx])
                                pu_pd_str = 'Reserved PU' if pwr_reserved_list[pwr_idx] else 'PU'
                                real_pu_pd_str_list.append(pu_pd_str)
                                real_power_str_list.append(pwr_rail_list[pwr_idx])
                    # 如果有pd
                    if gnd_rail_list:
                        for gnd_idx in range(len(gnd_rail_list)):
                            # 判断是否有电阻
                            if gnd_res_item_list[gnd_idx]:
                                # 如果有电阻则存入，否则不存入
                                real_resistance_str_list.append(gnd_res_item_list[gnd_idx])
                                pu_pd_str = 'Reserved PD' if gnd_reserved_list[gnd_idx] else 'PD'
                                real_pu_pd_str_list.append(pu_pd_str)
                                real_power_str_list.append(gnd_rail_list[gnd_idx])

                    # 存入所有情况
                    if real_resistance_str_list:
                        real_resistance_list.append('/'.join(real_resistance_str_list))
                        real_pu_pd_list.append('/'.join(real_pu_pd_str_list))
                        real_power_list.append('/'.join(real_power_str_list))
                        result_flag = False
                # 否则就直接存入
                else:
                    real_resistance_str_list = []
                    real_pu_pd_str_list = []
                    real_power_str_list = []
                    # 如果有pu
                    if pwr_rail_list:
                        for pwr_idx in range(len(pwr_rail_list)):
                            # 判断是否有电阻
                            if pwr_res_item_list[pwr_idx]:
                                # 如果有电阻则存入，否则不存入
                                real_resistance_str_list.append(pwr_res_item_list[pwr_idx])
                                pu_pd_str = 'Reserved PU' if pwr_reserved_list[pwr_idx] else 'PU'
                                real_pu_pd_str_list.append(pu_pd_str)
                                real_power_str_list.append(pwr_rail_list[pwr_idx])
                    # 如果有pd
                    if gnd_rail_list:
                        for gnd_idx in range(len(gnd_rail_list)):
                            # 判断是否有电阻
                            if gnd_res_item_list[gnd_idx]:
                                # 如果有电阻则存入，否则不存入
                                real_resistance_str_list.append(gnd_res_item_list[gnd_idx])
                                pu_pd_str = 'Reserved PD' if gnd_reserved_list[gnd_idx] else 'PD'
                                real_pu_pd_str_list.append(pu_pd_str)
                                real_power_str_list.append(gnd_rail_list[gnd_idx])

                    # 存入所有情况
                    if real_pu_pd_str_list:
                        real_resistance_list.append('/'.join(real_resistance_str_list))
                        real_pu_pd_list.append('/'.join(real_pu_pd_str_list))
                        real_power_list.append('/'.join(real_power_str_list))

                    else:
                        real_resistance_list.append(None)
                        real_pu_pd_list.append(None)
                        real_power_list.append(None)

                # ***************************** 进行两个版本之间的比较 *****************************
                # 将当前版本数据与common design进行比较
                real_net = real_signal_name_list[gpp_idx]
                real_pu_pd = real_pu_pd_list[gpp_idx]
                real_resistance = real_resistance_list[gpp_idx]
                real_voltage = real_power_list[gpp_idx]
                real_data_list = [real_net, real_pu_pd, real_resistance, real_voltage]
                # 比较当前版本与common design
                common_design_half_result_list = compare_with_common_design(result_flag, real_data_list,
                                                                            check_common_design_list,
                                                                            power_common_power_dict,
                                                                            [real_signal_name_list,
                                                                             real_pu_pd_list,
                                                                             real_resistance_list,
                                                                             real_power_list], gpp_idx)
                # 将real_power中的GND去掉
                for items_idx in range(len(real_power_list)):
                    items = real_power_list[items_idx]
                    if items:
                        items_list = items.split('/')
                        items_list = [item for item in items_list if item.find('GND') == -1]
                        real_power_list[items_idx] = '/'.join(items_list)

                # 和common design比较后真实数据已经经过了修改，因此重新赋值
                real_net = real_signal_name_list[gpp_idx]
                real_pu_pd = real_pu_pd_list[gpp_idx]
                real_resistance = real_resistance_list[gpp_idx]
                real_voltage = real_power_list[gpp_idx]
                real_data_list = [real_net, real_pu_pd, real_resistance, real_voltage]

                # 如果是aio的行，则直接忽略
                if gpp_idx in aio_ignore_idx_list:
                    result_list.append('1,1,1,1')
                else:
                    if first_version_flag is False:
                        bom_net = previous_design_pin_net_dict_bom[pin_name_item]
                        bom_pu_pd = previous_design_pin_pu_pd_dict_bom[pin_name_item]
                        bom_resistance = previous_design_pin_resistance_dict_bom[pin_name_item]
                        bom_voltage = previous_design_pin_voltage_dict_bom[pin_name_item]
                        bom_design_list = [bom_net, bom_pu_pd, bom_resistance, bom_voltage]

                        # 比较当前版本与前一版本
                        two_version_half_result_list_bom = compare_with_previous_design(real_data_list, bom_design_list,
                                                                                        real_signal_name_list, gpp_idx)

                        # 统一两个比较的结果
                        half_result_bom = ['1' if int(two_version_half_result_list_bom[i]) and
                                                  int(common_design_half_result_list[i]) else '0' for i in range(4)]
                        result_list.append(','.join(half_result_bom))
                    else:
                        result_list.append(','.join(common_design_half_result_list))
            else:
                # print(pin_name_item)
                # 如果是GROUP
                real_signal_name_list.append(None)
                real_resistance_list.append(None)
                real_pu_pd_list.append(None)
                real_power_list.append(None)
                error_message_list.append(None)
                result_list.append(None)
        # 将结果写入表格中并上色
        origin_excel_length = len(origin_excel_list[3])
        now_excel_length = int((origin_excel_length + len(result_list[1].split(','))) / 5)

        # 如果是第一版则直接生成标题
        if first_version_flag:
            worksheet.merge_range('B1:E1', 'Common Design', version_format)
            worksheet.merge_range('G1:J1', 'X00', version_format)
        # 如果不是第一版则写入原来的数据
        else:
            version_num = 0
            first_line_len = len(first_line_version_list)
            for first_idx in range(first_line_len):
                version_item = first_line_version_list[first_idx]
                if version_item:
                    if version_item != 'Common Design':
                        version_num += 1
                    col_1 = column_list[first_idx]
                    col_2 = column_list[first_idx + 3]
                    worksheet.merge_range('{}1:{}1'.format(col_1, col_2), version_item, version_format)
            col_3 = column_list[first_line_len]
            col_4 = column_list[first_line_len + 3]
            # 如果有新的
            if common_compare_diff_flag:
                # 写入现在的版本号
                col_5 = column_list[first_line_len + 5]
                col_6 = column_list[first_line_len + 8]
                worksheet.merge_range('{}1:{}1'.format(col_3, col_4), 'Common Design', version_format)
                worksheet.merge_range('{}1:{}1'.format(col_5, col_6), 'X{0:02d}'.format(version_num), version_format)
            else:
                worksheet.merge_range('{}1:{}1'.format(col_3, col_4), 'X{0:02d}'.format(version_num), version_format)

        if common_compare_diff_flag:
            worksheet.write_row('A2', first_title + title_list * (now_excel_length + 1), title_format)
        else:
            worksheet.write_row('A2', first_title + title_list * now_excel_length, title_format)
        # 写入原来的excel数据
        origin_row = 3
        for origin_excel in origin_excel_list:
            worksheet.write_row('A{}'.format(origin_row), origin_excel, column_pass_format)
            origin_row += 1

        # worksheet.write_row('A{}'.format(origin_row), origin_excel_list[-1], column_note_format)
        # 如果common design有变化则插入新的common design数据
        col = origin_excel_length
        common_design_list = [x[1:] for x in common_design_list[1:]]
        if common_compare_diff_flag:
            row = 3
            for com_item in common_design_list:
                row += 1
                worksheet.write_row('{}{}'.format(column_list[col], row), com_item[:-1], column_pass_format)
                # note单独存入
                worksheet.write('{}{}'.format(column_list[col + 4], row), com_item[-1], column_note_format)

            col = origin_excel_length + 5
        # 写入新的excel数据，并上色
        row = 2
        for idx in range(len(real_signal_name_list)):
            row += 1
            col_new = col
            # 跳过group分割
            if result_list[idx]:
                input_result_list = result_list[idx].split(',')
                input_signal_name_list = real_signal_name_list[idx]
                if input_signal_name_list:
                    input_signal_name = input_signal_name_list[0]
                else:
                    input_signal_name = input_signal_name_list
                input_pu_pd = real_pu_pd_list[idx]
                input_resistance = real_resistance_list[idx]
                input_power_rail = real_power_list[idx]
                # 写入并上色
                if input_result_list[0] == '1':
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_signal_name, column_pass_format)
                else:
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_signal_name, column_fail_format)
                col_new += 1

                if input_result_list[1] == '1':
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_pu_pd, column_pass_format)
                else:
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_pu_pd, column_fail_format)
                col_new += 1

                if input_result_list[2] == '1':
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_resistance, column_pass_format)
                else:
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_resistance, column_fail_format)
                col_new += 1

                if input_result_list[3] == '1':
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_power_rail, column_pass_format)
                else:
                    worksheet.write('{}{}'.format(column_list[col_new], row), input_power_rail, column_fail_format)
        # 给pin name上色
        pin_name_list = [x[0] for x in origin_excel_list[1:]]
        worksheet.write_column('A4', pin_name_list, column_pin_name_format)

        # 给之前note上色
        for i in range(col_idx):
            note_count = (i + 1) * 5
            common_note_list = [x[note_count] for x in origin_excel_list[1:]]
            worksheet.write_column('{}4'.format(column_list[note_count]), common_note_list, column_note_format)
        # 给新生成的note上色
        note_list = len(common_note_list) * [None]
        if common_compare_diff_flag:
            worksheet.write_column('{}4'.format(column_list[(col_idx + 2) * 5]), note_list, column_note_format)
        else:
            worksheet.write_column('{}4'.format(column_list[(col_idx + 1) * 5]), note_list, column_note_format)

        # 给group上色
        if common_compare_diff_flag:
            group_lists = [x + [None] * 10 for x in origin_excel_list if x[0].upper().find('GROUP') > -1]
        else:
            group_lists = [x + [None] * 5 for x in origin_excel_list if x[0].upper().find('GROUP') > -1]

        for group_x in range(len(group_idx_list)):
            worksheet.write_row('A{}'.format(group_idx_list[group_x] + 3), group_lists[group_x], column_group_format)

        # 将之前数据错误的变成红色
        if not first_version_flag:
            open_book = openpyxl.load_workbook(gpio_version_input_excel_path)  # 读取excel
            open_sheet = open_book.get_sheet_by_name(sheet_name)  # 读取Sheet
            rows, cols = open_sheet.max_row, open_sheet.max_column
            for i in range(1, rows):
                for j in range(1, cols):
                    ce = open_sheet.cell(row=i, column=j)
                    fill = ce.fill
                    if fill.start_color.rgb == 'FFFC2443':
                        worksheet.write(i - 1, j - 1, ce.value, column_fail_format)

        # 设置冻结窗口
        worksheet.freeze_panes(2, 1)

    workbook.close()

    # 设置自适应
    sheets_autofit(output_excel_path)


def main():
    generate_report()


if __name__ == '__main__':
    main()
